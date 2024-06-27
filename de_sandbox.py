from pathlib import Path
import openpyxl
import openpyxl.workbook
import openpyxl.worksheet
import regex 
from typing import List
import json


def load_workbook(filepath: Path) -> openpyxl.workbook.workbook.Workbook:
    # if filepath.suffix != ".xlsx":
    #     raise ValueError(
    #         f"The inputted file is of type {filepath.suffix}. Please provide a .xlsx file."
    #     )

    try:
        workbook = openpyxl.load_workbook(filepath)

    except Exception as e:
        raise f"An unexpected error occurred: {e}"

    else:
        return workbook
    

def load_worksheet(
    workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' cannot be found in workbook.")

    else:
        return workbook[sheet_name]
    

def parse_sheet_data(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> List[List]:
    tables = []
    current_table = []
    for row in worksheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            if current_table:
                tables.append(current_table)
                current_table = []

        else:
            current_table.append([*row])

    if current_table:
        tables.append(current_table)

    return tables



filepath = Path.cwd() / "survey_pk" / "Database and questions.xlsx"

# Compile a RegEx pattern to identify screening question.
#   (?<=^\[|^)? ensures that the match starts with either "[" or "", but does not match this part.
#   S\d+ quite literally matches S, followed by 1 or more digits
#   (?=[^\]:]*[\]:]) ensures the "S\d+" is followed by any character that isn't "]" or ":", and ends with an optional "]" followed by a necessary ":"

# re_tests = """[S1r4oe]: Which geography do you cover? - Other (please specify)
# [Q0]: Please choose your language preference:
# [HidSamp]: Hidden to punch sample
# [S1]: Which geography do you cover?
# S1_term: S0 TERMINATE
# [noooo S1]
# S1 Please NO""".split("\n")

# for st in re_tests:
#     print(st, regex.search(question_regex, st))
#     # print(st, regex.search(r'(?<=^\[?)S\d+(?=[^\]:]*\]?:)', st))


question_regex = regex.compile(r'(?<=^\[?)(S|Q)\d+(?=[^\]:]*\]?:)')
all_field_regex = regex.compile(r'(?<=^)\[?[\w^\]^:]+\]?:')


wb = load_workbook(filepath=filepath)
sheet = load_worksheet(workbook=wb, sheet_name="Datamap")


def clean_row(row):
    cleaned_row = []
    for cell in row:
        if cell.value == "":
            cell.value = None


def get_next_row_safe(row_generator):
    for row in row_generator:
        # BUG: I don't know why this didn't work previously.
        clean_row(row)
        first_cell, *_ = row
        if all(c.value is None for c in row):
            yield False

        elif first_cell.value and regex.search(all_field_regex, first_cell.value):
            yield False


        else:
            yield row


# metadata = dict()
# screening_questions = dict()
# survey_questions = dict()
all_keys = dict()


prev_row = tuple()
for row in sheet.iter_rows():
    clean_row(row)
    if not all(c.value is None for c in row):
        first_cell, *_ = row
        assert first_cell.column == 1
        if all(c.value is None for c in prev_row):
            # Check that previous row was empty
            
            # Check to see if we have a question/screening field
            field_code = regex.search(all_field_regex, first_cell.value)
            if not field_code:
                print(f"Could not identify a field code for the cell {first_cell.value} ({first_cell.row}, {first_cell.column})")
                continue

            else:
                field_code = field_code.group(0)
            
            is_column = regex.search(r"^\[\w+\]:$", field_code) is not None
            if is_column:
                key = field_code[1:-2]

            else:
                key = field_code

            field_search = regex.search(question_regex, field_code)
            is_metadata = regex.search(question_regex, field_code) is None

            # if field_search:
            #     key = field_search.group(0)
            #     is_metadata = False

            # else:
            #     is_metadata = True
            #     if is_column:
            #         key = field_code[1:-2]
            #     else:
            #         key = field_code
            
            if key in all_keys.keys():
                print(f"Found a duplicate key {key} for {first_cell.value}")

            else:
                primary_text = regex.search(r'(?<=^\[?\w+\]?: ).+', first_cell.value)
                if not primary_text:
                    print(f"No primary text found for {field_code}")
                
                else:
                    primary_text = primary_text.group(0)

                all_keys[key] = {
                    "is_column": is_column,
                    "is_metadata": is_metadata,
                    "primary_text": primary_text,
                    "schema_row_start": first_cell.row,
                }

            curr_key = key
        
        else:
            if first_cell.row == all_keys[curr_key]["schema_row_start"] + 1:
                field_type = regex.search(r"(?<=Values: ?)\d+-\d+$", first_cell.value)
                if field_type:
                    value_map_start = field_type.group(0).split("-")[0]
                    value_map_end = field_type.group(0).split("-")[1]
                    
                    if str.isnumeric(value_map_start) and str.isnumeric(value_map_end):
                        encoding_start = first_cell.row + 1
                        encoding_end = first_cell.row + (int(value_map_end) - int(value_map_start)) + 1
                        encoding_table = [
                            [c.value for c in r]
                            for r in sheet.iter_rows(min_col=2, max_col=3, min_row=encoding_start, max_row=encoding_end)
                        ]
                        all_keys[key]["encodings"] = dict(encoding_table)

                    else:
                        print(f"Found Non-Numeric Value Map {field_type.group(0)} for {key}. Skipping.")
                        continue

                if all_keys[curr_key]["is_column"]:
                    all_keys[curr_key]["related_columns"] = [curr_key]

                else:
                    sub_row_start = encoding_end + 1
                    column_rows = dict()
                    next_row_iter = get_next_row_safe(sheet.iter_rows(min_row=sub_row_start))#, max_row=all_keys[curr_key]["schema_row_end"]))
                    next_row = next(next_row_iter)
                    while next_row:
                        non_null_row = [c.value for c in next_row if c.value]
                        if len(non_null_row) != 2:
                            print(f"Row encoding error for {curr_key} row ({sub_row_start}). Row table contains more than two variables: {non_null_row}.")

                        else:
                            row_id, row_text = non_null_row
                            column_rows[row_id] = {
                                "row_number": row_id.rsplit("r", 1)[-1],
                                "row_text": row_text
                            }
                        
                        next_row = next(next_row_iter)                        

                    all_keys[curr_key]["related_columns"] = list(column_rows.keys())
                    all_keys[curr_key]["rows"] = column_rows
                        

            # if not regex_search:
            #     # This is a non-screener or survey question I.E. Metadata
            #     if first_cell.value in all_keys.keys():
            #         print(f"Found a duplicate metadata key at row {first_cell.row}: {first_cell.value}")

            #     else:
            #         key = regex.search(all_field_regex, first_cell.value).group(0)
            #         all_keys[regex.search(all_field_regex, first_cell.value).group(0)] = {
            #             "schema_row_start": first_cell.row
            #         }

            # else:
            #     # print(first_cell.value, regex_search)
            #     primary_key = regex_search.group(0)
            #     if primary_key in all_keys.keys():
            #         print(f"Found a duplicate survey/screening key at row {first_cell.row}: {primary_key} + {first_cell.value}")

            #     else:
            #         all_keys[primary_key] = {
            #             "schema_row_start": first_cell.row
            #         }
        
        # else:
        #     pass

    prev_row = row

# Sort the dictionary items based on 'schema_row_start' and iterate through them
sorted_keys = sorted(all_keys, key=lambda x: all_keys[x]["schema_row_start"])

# Loop through the sorted keys to set 'schema_row_end' for each
for i in range(len(sorted_keys) - 1):
    current_key = sorted_keys[i]
    next_key = sorted_keys[i + 1]
    # Set 'schema_row_end' to be one less than the 'schema_row_start' of the next item
    all_keys[current_key]['schema_row_end'] = all_keys[next_key]["schema_row_start"] - 1

# Add max row at the bottom
last_key = sorted_keys[-1]
all_keys[last_key]["schema_row_end"] = sheet.max_row

print(json.dumps(all_keys, indent=4))


# def get_next_row_safe(row_generator):
#     for row in row_generator:
#         first_cell, *_ = row
#         if all(c.value is None for c in row):
#             yield False

#         elif first_cell.value and regex.search(all_field_regex, first_cell.value):
#             yield False

#         else:
#             yield row


# for primary_key, table_data in all_keys.items():
#     row_start = table_data["schema_row_start"]
#     row_end = table_data["schema_row_end"]
#     primary_text = regex.search(
#         r"(?<=: )\w+", sheet.cell(row=row_start, column=1).value
#     )
#     is_col = table_data["is_column"]
#     field_type = sheet.cell(row=row_start+1, column=1).value
#     sub_dict = dict()
#     prev_row = (None)
#     map_range = regex.search(r"(?<=Values: ?)\d+-\d+$", field_type)
#     table_rows = sheet.iter_rows(min_row=row_start, max_row=row_end)
#     sub_tables = dict()

#     if map_range:
        # value_map_start = map_range.group(0).split("-")[0]
        # value_map_end = map_range.group(0).split("-")[1]
        
        # if str.isnumeric(value_map_start) and str.isnumeric(value_map_end):
        #     encoding_start = row_start + 2
        #     encoding_end = row_start + (int(value_map_end) - int(value_map_start)) + 2
        #     encoding_table = [
        #         [c.value for c in r]
        #         for r in sheet.iter_rows(min_col=2, max_col=3, min_row=encoding_start, max_row=encoding_end)
        #     ]
        #     all_keys[primary_key]["encodings"] = dict(encoding_table)

        # else:
        #     print(f"Found Non-Numeric Value Map {map_range} for {primary_key}. Skipping.")
        #     continue

#         if is_col:
#             all_keys[primary_key]["column_type"] = "single_select"
#             all_keys[primary_key]["related_columns"] = [primary_key]

#         else:
            # sub_row_start = encoding_end + 1
            # column_rows = []
            # next_row_iter = get_next_row_safe(sheet.iter_rows(min_row=sub_row_start, max_row=row_end))
            # next_row = next(next_row_iter)
            # while next_row:
            #     column_rows.append([c.value for c in next_row if c.value])
            #     next_row = next(next_row_iter)

            # print(primary_key, json.dumps(column_rows, indent=4))


#     else:
#         # print(f"Found No Value Map of {field_type} for {primary_key}")
#         all_keys[primary_key]["column_type"] = field_type
#         if not is_col:
#             print(f"Warning. Primary Key {primary_key} is not a column and of unknown type {field_type}.")

#         else:
#             all_keys[primary_key]["related_columns"] = [primary_key]
        