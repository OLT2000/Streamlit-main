import openpyxl
import json
from pathlib import Path
from typing import List

import openpyxl.workbook


def load_workbook(filepath: Path) -> openpyxl.workbook.workbook.Workbook:
    if filepath.suffix != ".xlsx":
        raise ValueError(
            f"The inputted file is of type {filepath.suffix}. Please provide a .xlsx file."
        )

    try:
        workbook = openpyxl.load_workbook(filepath)

    except Exception as e:
        raise f"An unexpected error occurred: {e}"

    else:
        return workbook


def load_worksheet(
    workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' cannot be found in workbook.")

    else:
        return workbook[sheet_name]


def parse_sheet_data(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> List[List]:
    tables = []
    current_table = []
    for row in worksheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            if current_table:
                tables.append(current_table)
                current_table = []

        else:
            current_table.append([r for r in row if r is not None])

    if current_table:
        tables.append(current_table)

    return tables


def parse_table_data(table: List[List]) -> dict:
    if len(table) < 2:
        return None

    column_name, column_description = table.pop(0)
    column_type = table.pop(0)[1]
    encodings = {key: value for key, value in table if key is not None}

    return {
        "column_name": column_name,
        "column_description": column_description,
        "column_type": column_type,
        "encodings": encodings,
    }


def load_question_schema(filepath: Path, sheet_name: str) -> List[dict]:
    wb = load_workbook(filepath=filepath)
    sheet = load_worksheet(workbook=wb, sheet_name=sheet_name)
    processed_tables = parse_sheet_data(sheet)

    return [parse_table_data(t) for t in processed_tables]


# def read_excel_sheet(file_path: Path, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.get(sheet_name)
#     return sheet


# def parse_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List:
#     tables = []
#     current_table = []
#     for row in sheet.iter_rows(values_only=True):
#         if all(cell is None for cell in row):
#             if current_table:
#                 tables.append(current_table)
#                 current_table = []

#         else:
#             current_table.append(row)

#     if current_table:
#         tables.append(current_table)

#     return tables


# def parse_table(table):
#     if len(table) < 2:
#         return None
#     column_name = table[0][0]
#     column_description = table[0][1]
#     column_type = table[1][1]
#     encodings = {}
#     for row in table[2:]:
#         if row[0] is not None:
#             encodings[row[0]] = row[1]

#     return {
#         "column_name": column_name,
#         "column_description": column_description,
#         "column_type": column_type,
#         "encodings": encodings
#     }


# def convert_to_json(tables):
#     json_output = []
#     for table in tables:
#         parsed_table = parse_table(table)
#         if parsed_table:
#             json_output.append(parsed_table)
#     return json.dumps(json_output, indent=4)


# def main(file_path):
#     sheet = read_excel_sheet(file_path, sheet_name="Questions")
#     tables = parse_sheet(sheet)
#     json_data = convert_to_json(tables)
#     print(json_data)


if __name__ == "__main__":
    import pandas as pd
    file_path = Path(__file__).parent.parent / "survey.xlsx"
    json_output = load_question_schema(filepath=file_path, sheet_name="Questions")
    question_df = pd.DataFrame.from_records(json_output).drop(columns=["encodings"])

    def str_helper(row):
        if row["column_type"] == "checkbox" and "_" in row["column_name"]:
            output = row["column_name"].split("_", 1)
        
        else:
            output = [row["column_name"], ""]
        
        return output
        
    question_df[["qcode", "subq"]] = question_df.apply(lambda r: str_helper(r), axis=1, result_type="expand")
    with pd.option_context({"display.max_columns": None}):
        print(question_df.head(20))
    # print(json.dumps())
    # with open("question_schema.json", "w") as f:
    #     json.dump(json_output, f, indent=4)


    # data = pd.read_excel(file_path, sheet_name="Data", header=0)
    # with open("../survey_data.csv", "w") as f:
    #     data.to_csv(f, header=True, index=False)
