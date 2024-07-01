
import openpyxl
import json
from pathlib import Path
from typing import List
import openpyxl.workbook
import regex

import openpyxl.worksheet
import openpyxl.worksheet.worksheet


# Define our Regular Expressions
identifier_regex = regex.compile(r"(?<=^)\[?[\w^\]^:]+\]?:")
non_metadata_regex = regex.compile(r'(?<=^)(S|Q)\d+(?=\w*)')
question_text_regex = regex.compile(r"(?<=\[?\w+\]?: ).+")
value_match_regex = regex.compile(r"(?<=Values: ?)\d+-\d+$")


def load_workbook(filepath: Path) -> openpyxl.workbook.workbook.Workbook:
    # if filepath.suffix != ".xlsx":
    #     raise ValueError(
    #         f"The inputted file is of type {filepath.suffix}. Please provide a .xlsx file."
    #     )

    try:
        workbook = openpyxl.load_workbook(filepath)

    except Exception as e:
        raise f"An unexpected error occurred: {e}"

    else:
        return workbook


def load_worksheet(
    workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' cannot be found in workbook.")

    else:
        return workbook[sheet_name]


def row_iterator(row_generator):
    for row in row_generator:
        clean_excel_row(row)
        first_cell, *_ = row
        if all(c.value is None for c in row):
            yield False

        elif first_cell.value and regex.search(identifier_regex, first_cell.value):
            yield False

        else:
            yield row


def clean_excel_row(row: tuple) -> None:
    for cell in row:
        if cell.value == "":
            cell.value = None


def process_atheneum_schema(excel_file, sheet_name = "Datamap"):
# def process_atheneum_schema(schema_sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    wb = load_workbook(excel_file)
    schema_sheet = load_worksheet(workbook=wb, sheet_name=sheet_name)

    cleaned_keys: dict = dict()
    previous_row = tuple()

    # Loop through the sheet rows
    for row in schema_sheet.iter_rows():
        clean_excel_row(row)
        first_cell, *_ = row
        assert first_cell.column == 1
        # If the previous row is empty, we have started a new table
        if all(cell.value is None for cell in previous_row):
            # Extract the table ID (I.E. The first element of the sub-table)
            identifier_match = regex.search(identifier_regex, first_cell.value)
            if not identifier_match:
                print(f"Could not identify an identifier for the value {first_cell.value} in position ({first_cell.row}, {first_cell.column}). Skipping.")
                continue

            else:
                table_id = identifier_match.group(0)

            is_column: bool = regex.search(r"^\[\w+\]:$", table_id) is not None
            if is_column:
                table_key = table_id[1:-2]

            else:
                table_key = table_id

            if table_key in cleaned_keys:
                print(f"Found a duplicate table key {table_key}. Skipping.")
                continue

            else:
                question_text = regex.search(question_text_regex, first_cell.value)

                if not question_text:
                    print(f"No question text found for table key {table_key}.")
                    question_text = ""
                
                else:
                    question_text = question_text.group(0)

                cleaned_keys[table_key] = {
                    "question_text": question_text,
                    "is_column": is_column,
                    "is_metadata": regex.search(non_metadata_regex, table_key) is None,
                    "schema_row_start": first_cell.row
                }
            
            current_key = table_key
        
        else:
            # The cell below the table key should be the encoding/field type
            if first_cell.row == cleaned_keys[current_key]["schema_row_start"] + 1:
                field_type = regex.search(value_match_regex, first_cell.value)
                if field_type:
                    value_map_range = field_type.group(0).split("-")
                    if str.isnumeric(value_map_range[0]) and str.isnumeric(value_map_range[1]):
                        encoding_start_row = first_cell.row + 1
                        encoding_end_row = encoding_start_row + (int(value_map_range[1]) - int(value_map_range[0]))

                        # Process the encodings table
                        encodings_table = [
                            [str(cell.value) for cell in code_row]
                            for code_row in schema_sheet.iter_rows(min_col=2, max_col=3, min_row=encoding_start_row, max_row=encoding_end_row)
                        ]

                        cleaned_keys[current_key]["encodings"] = dict(encodings_table)

                    else:
                        print(f"Found Non-Numeric Value Map '{field_type.group(0)}' for {current_key}. Skipping.")
                        continue

                if cleaned_keys[current_key]["is_column"]:
                    cleaned_keys[current_key]["related_columns"] = [current_key]
                    # cleaned_keys[current_key][]

                else:
                    row_definition_start = encoding_end_row + 1
                    row_iter_tool = row_iterator(schema_sheet.iter_rows(min_row=row_definition_start))
                    next_row = next(row_iter_tool)
                    question_rows = dict()
                    while next_row:
                        non_null_row = [cell.value for cell in next_row if cell.value]
                        if len(non_null_row) == 1:
                            print(f"Could not find a valid encoding, defaulting to [{non_null_row[0]}, None]")
                            non_null_row.append(None)

                        elif len(non_null_row) != 2:
                            print(f"Row encoding error for {current_key} at row {row_definition_start}. The values are: {non_null_row}")

                        else:
                            row_id, row_text = non_null_row
                            question_rows[row_id[1:-1]] = {
                                "row_number": row_id.rsplit("r", 1)[-1],
                                "row_text": row_text
                            }
                        
                        next_row = next(row_iter_tool)
                    
                    cleaned_keys[current_key]["related_columns"] = list(question_rows.keys())
                    cleaned_keys[current_key]["rows"] = question_rows

        previous_row = row

    return cleaned_keys


if __name__ == "__main__":
    filepath = Path.cwd() / "survey_pk" / "Database and questions.xlsx"
    # wb = load_workbook(filepath=filepath)
    # sheet = load_worksheet(workbook=wb, sheet_name="Datamap")

    test_keys = process_atheneum_schema(filepath)
    print(json.dumps(test_keys, indent=4))